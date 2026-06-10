"""
Excel 报告生成器 — 复用 22iterate_evalv3.py 的报告模式。

生成：
  - Sheet 1 "详细结果": 逐条任务结果
  - Sheet 2 "统计汇总": 每个组合的指标统计
  - Sheet 3 "模型对比": 跨模型对比表
"""
from __future__ import annotations

import os
from dataclasses import asdict
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..core.trajectory import ComboResult, ExperimentResult, Trajectory


# ── 样式常量 ──
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
GREEN_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="4472C4")
HEADER_FONT_WHITE = Font(bold=True, color="FFFFFF")


class ExcelWriter:
    """Excel 报告生成器。"""

    def __init__(self, output_dir: str):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def write(self, experiment_result: ExperimentResult, env_name: str = "") -> str:
        """生成完整 Excel 报告。

        Returns:
            生成的文件路径。
        """
        wb = openpyxl.Workbook()

        # Sheet 1: 详细结果
        ws_detail = wb.active
        ws_detail.title = "详细结果"
        self._write_detail_sheet(ws_detail, experiment_result)

        # Sheet 2: 统计汇总
        ws_stats = wb.create_sheet("统计汇总")
        self._write_stats_sheet(ws_stats, experiment_result)

        # Sheet 3: 模型对比
        ws_compare = wb.create_sheet("模型对比")
        self._write_compare_sheet(ws_compare, experiment_result)

        # 保存
        stem = f"{experiment_result.experiment_name}_{env_name}" if env_name else experiment_result.experiment_name
        filename = f"{stem}_report.xlsx"
        filepath = self.output_dir / filename
        wb.save(str(filepath))
        print(f"  Excel 报告已保存: {filepath}")
        return str(filepath)

    def _write_detail_sheet(
        self, ws: openpyxl.worksheet.worksheet.Worksheet,
        experiment_result: ExperimentResult,
    ) -> None:
        """写入详细结果 Sheet。"""
        # 表头
        headers = [
            "LLM", "Harness", "Environment",
            "任务ID", "状态",
            "最终答案", "Ground Truth",
            "部分准确率", "格式合规",
            "总Token", "延迟(ms)",
            "错误字段",
        ]
        ws.append(headers)
        self._style_header(ws, len(headers))

        # 数据行
        for traj in experiment_result.get_all_trajectories():
            diff_fields = traj.metadata.get("field_results", {})
            wrong_fields = [k for k, v in diff_fields.items() if not v]

            row = [
                traj.llm_name,
                traj.harness_name,
                traj.env_name,
                traj.task_id,
                traj.status,
                traj.final_answer[:200] if traj.final_answer else "",
                str(traj.ground_truth)[:200] if traj.ground_truth else "",
                f"{traj.scores.get('partial_completion', 0):.1%}",
                "✓" if traj.scores.get("format_compliance", 0) > 0 else "✗",
                traj.total_input_tokens + traj.total_output_tokens,
                f"{traj.total_latency_ms:.0f}",
                ", ".join(wrong_fields) if wrong_fields else "",
            ]
            ws.append(row)

            # 错误行标红
            if traj.status != "success":
                row_idx = ws.max_row
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = YELLOW_FILL

        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        ws.freeze_panes = "A2"

    def _write_stats_sheet(
        self, ws: openpyxl.worksheet.worksheet.Worksheet,
        experiment_result: ExperimentResult,
    ) -> None:
        """写入统计汇总 Sheet。"""
        headers = [
            "LLM", "Harness", "Environment",
            "总任务数", "成功数", "部分成功数", "失败数", "超时数", "错误数",
            "成功率", "平均得分", "格式合规率",
            "平均延迟(ms)", "平均Token",
        ]
        ws.append(headers)
        self._style_header(ws, len(headers))

        for combo in experiment_result.combo_results:
            s = combo.summary
            if not s:
                continue
            row = [
                combo.llm_name,
                combo.harness_name,
                combo.env_name,
                s.get("total_tasks", 0),
                s.get("success_count", 0),
                s.get("partial_count", 0),
                s.get("failure_count", 0),
                s.get("timeout_count", 0),
                s.get("error_count", 0),
                f"{s.get('success_rate', 0):.1%}",
                f"{s.get('avg_score', 0):.1%}",
                f"{s.get('format_compliance_rate', 0):.1%}",
                f"{s.get('avg_latency_ms', 0):.0f}",
                f"{s.get('avg_tokens', 0):.0f}",
            ]
            ws.append(row)

        ws.freeze_panes = "A2"

    def _write_compare_sheet(
        self, ws: openpyxl.worksheet.worksheet.Worksheet,
        experiment_result: ExperimentResult,
    ) -> None:
        """写入模型对比 Sheet（LLM × Harness 成功率热力图）。"""
        matrix = experiment_result.get_comparison_matrix()

        # 表头
        all_harnesses = sorted(set(
            h for llm_data in matrix.values() for h in llm_data.keys()
        ))
        headers = ["LLM \\ Harness"] + all_harnesses
        ws.append(headers)
        self._style_header(ws, len(headers))

        for llm_name in sorted(matrix.keys()):
            row = [llm_name]
            for h in all_harnesses:
                data = matrix[llm_name].get(h, {})
                rate = data.get("success_rate", 0)
                cell_val = f"{rate:.1%}"
                row.append(cell_val)
            ws.append(row)

            # 颜色编码
            row_idx = ws.max_row
            for col_idx, h in enumerate(all_harnesses, 2):
                data = matrix[llm_name].get(h, {})
                rate = data.get("success_rate", 0)
                cell = ws.cell(row=row_idx, column=col_idx)
                if rate >= 0.8:
                    cell.fill = GREEN_FILL
                elif rate < 0.5:
                    cell.fill = RED_FILL

        ws.freeze_panes = "B2"

    def _style_header(
        self, ws: openpyxl.worksheet.worksheet.Worksheet,
        num_cols: int,
    ) -> None:
        """设置表头样式。"""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
